"""
分析Excel的Q1税收计算
"""
import openpyxl

TEMPLATE_PATH = r"d:\work\110agenticAI\code\ai-deal-associate\backend\data\templates\MS Canopy Template -v5.xlsx"

wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
cf_sheet = wb["Cash Flows"]

print("="*70)
print("分析Excel Q0和Q1现金流组成")
print("="*70)

# 检查Q0 (Col 20) 的所有组成部分
print("\n📋 Q0 (Col 20) 现金流组成:")
print("-"*50)
key_rows = {
    23: "GRI",
    24: "Rent Free/Turnover",
    26: "OpEx",
    27: "EBITDA",
    30: "Interest",
    35: "Tax",
    38: "Cash Flow Operations",
    56: "Gross Disposal Proceeds",
    59: "Net Disposal Proceeds",
    61: "Purchase Price",
    62: "Transfer Tax",
    63: "Closing Costs",
    64: "Total Acquisition Price",
    91: "CF Ops",
    94: "CF Investing",
    96: "Debt Issuance",
    97: "Debt Repayment",
    99: "CF Financing",
    102: "Levered Post Tax CF"
}

for row, name in key_rows.items():
    value_q0 = cf_sheet.cell(row=row, column=20).value
    value_q1 = cf_sheet.cell(row=row, column=21).value
    if value_q0 is not None or value_q1 is not None:
        v0 = f"€{value_q0:,.2f}" if value_q0 else "-"
        v1 = f"€{value_q1:,.2f}" if value_q1 else "-"
        print(f"Row {row:3d} ({name:25s}): Q0={v0:>20s}  Q1={v1:>15s}")

# 检查Q1的税收计算
print("\n📋 税收计算分析:")
print("-"*50)
q1_ebitda = cf_sheet.cell(row=27, column=21).value
q1_interest = cf_sheet.cell(row=30, column=21).value
q1_tax = cf_sheet.cell(row=35, column=21).value
q1_cf_ops = cf_sheet.cell(row=38, column=21).value

print(f"Q1 EBITDA: €{q1_ebitda:,.2f}" if q1_ebitda else "Q1 EBITDA: -")
print(f"Q1 Interest: €{q1_interest:,.2f}" if q1_interest else "Q1 Interest: -")
print(f"Q1 Tax: €{q1_tax:,.2f}" if q1_tax else "Q1 Tax: -")
print(f"Q1 CF Ops: €{q1_cf_ops:,.2f}" if q1_cf_ops else "Q1 CF Ops: -")

# 检查Python的计算
print("\n📋 Python计算对比:")
print("-"*50)
quarterly_ebitda = 102372.75
quarterly_interest = 54336.31
pre_tax_cf = quarterly_ebitda - quarterly_interest
python_tax = pre_tax_cf * 0.25
python_cf = pre_tax_cf - python_tax

print(f"Python EBITDA: €{quarterly_ebitda:,.2f}")
print(f"Python Interest: €{quarterly_interest:,.2f}")
print(f"Python Pre-tax CF: €{pre_tax_cf:,.2f}")
print(f"Python Tax (25%): €{python_tax:,.2f}")
print(f"Python CF: €{python_cf:,.2f}")

if q1_ebitda and q1_interest:
    excel_pre_tax = q1_ebitda - abs(q1_interest) if q1_interest < 0 else q1_ebitda - q1_interest
    print(f"\nExcel Pre-tax CF (EBITDA - Interest): €{excel_pre_tax:,.2f}")
    if q1_tax:
        print(f"Excel Tax: €{q1_tax:,.2f}")
        implied_rate = q1_tax / excel_pre_tax if excel_pre_tax > 0 else 0
        print(f"Implied Tax Rate: {implied_rate*100:.2f}%")

# 检查Transfer Tax和Closing Costs
print("\n📋 Q0 Transfer Tax和Closing Costs:")
transfer_tax = cf_sheet.cell(row=62, column=20).value
closing_costs = cf_sheet.cell(row=63, column=20).value
total_acq = cf_sheet.cell(row=64, column=20).value

print(f"Transfer Tax (Row 62): €{transfer_tax:,.2f}" if transfer_tax else "Transfer Tax: -")
print(f"Closing Costs (Row 63): €{closing_costs:,.2f}" if closing_costs else "Closing Costs: -")
print(f"Total Acq Price (Row 64): €{total_acq:,.2f}" if total_acq else "Total Acq: -")

# Purchase Price验证
purchase_price = cf_sheet.cell(row=61, column=20).value
print(f"\nPurchase Price (Row 61): €{purchase_price:,.2f}" if purchase_price else "Purchase Price: -")

if purchase_price and transfer_tax and closing_costs:
    expected_total = purchase_price + transfer_tax + closing_costs
    print(f"Expected Total: €{expected_total:,.2f}")

wb.close()
