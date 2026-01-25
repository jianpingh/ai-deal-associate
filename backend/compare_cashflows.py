"""
详细对比Excel和Python的现金流序列
"""
import openpyxl
import sys
sys.path.insert(0, 'd:/work/110agenticAI/code/ai-deal-associate/backend')

from deal_agent.nodes.model import calculate_metrics_excel_compatible, xirr
from datetime import datetime

TEMPLATE_PATH = r"d:\work\110agenticAI\code\ai-deal-associate\backend\data\templates\MS Canopy Template -v5.xlsx"

wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
cf_sheet = wb["Cash Flows"]

print("="*70)
print("Excel vs Python 现金流对比")
print("="*70)

# 读取Excel的实际现金流 (Row 102, Col 20 onwards)
excel_cfs = []
excel_dates = []
for col_idx in range(20, 60):
    cf = cf_sheet.cell(row=102, column=col_idx).value
    dt = cf_sheet.cell(row=78, column=col_idx).value
    if cf is not None:
        excel_cfs.append(cf)
        excel_dates.append(dt)
    else:
        break

print(f"\n📊 Excel现金流 (共{len(excel_cfs)}期):")
for i, (cf, dt) in enumerate(zip(excel_cfs, excel_dates)):
    if i == 0 or i == len(excel_cfs)-1 or cf != excel_cfs[1]:
        print(f"  Q{i}: {dt.strftime('%Y-%m-%d') if dt else 'N/A'} = €{cf:,.2f}")
print(f"  ... (Q1-Q{len(excel_cfs)-2} each = €{excel_cfs[1]:,.2f})")

# 计算Python现金流
inputs = {
    "entry_yield": 0.065,
    "exit_yield": 0.0475,
    "rent_growth": 0.02,
    "ltv": 0.60,
    "interest_rate": 0.0575,
    "purchasers_costs": 0.065,
    "hold_period": 7,
    "opex_ratio": 0.0,
}
correct_annual_noi = 409_491

result = calculate_metrics_excel_compatible(inputs, passing_rent_total=correct_annual_noi)

# 计算Excel的XIRR和Multiple
print("\n📊 Excel现金流验证:")
print(f"  Q0 (initial): €{excel_cfs[0]:,.2f}")
print(f"  Q1-Q27 (operations): €{excel_cfs[1]:,.2f} each")
print(f"  Q28 (exit): €{excel_cfs[28]:,.2f}")

# 计算XIRR
try:
    excel_irr = xirr(excel_cfs, excel_dates)
    print(f"\n📈 Excel XIRR计算结果: {excel_irr*100:.2f}%")
except Exception as e:
    print(f"XIRR计算失败: {e}")

# Multiple计算
positive_cf = sum(cf for cf in excel_cfs if cf > 0)
negative_cf = sum(cf for cf in excel_cfs if cf < 0)
excel_multiple = -positive_cf / negative_cf
print(f"📈 Excel Multiple计算结果: {excel_multiple:.2f}x")

# 对比
print("\n📊 对比:")
print(f"  Q0 差异: €{excel_cfs[0] - (-3033383.33):,.2f}")
print(f"  Q1 差异: €{excel_cfs[1] - 36027:,.2f}")
print(f"  Exit差异: €{excel_cfs[28] - 6010232:,.2f}")

# 详细分析Excel的Q0组成
print("\n📋 Excel Q0组成分析 (Row 93-101):")
for row in range(93, 103):
    label = cf_sheet.cell(row=row, column=3).value or cf_sheet.cell(row=row, column=2).value
    value = cf_sheet.cell(row=row, column=20).value
    if value:
        print(f"  Row {row}: {label} = €{value:,.2f}")

# 详细分析Excel的Q1组成
print("\n📋 Excel Q1组成分析 (Row 93-101):")
for row in range(93, 103):
    label = cf_sheet.cell(row=row, column=3).value or cf_sheet.cell(row=row, column=2).value
    value = cf_sheet.cell(row=row, column=21).value
    if value:
        print(f"  Row {row}: {label} = €{value:,.2f}")

# 详细分析Excel的Q28组成
print("\n📋 Excel Q28 (Exit)组成分析 (Row 93-101):")
for row in range(93, 103):
    label = cf_sheet.cell(row=row, column=3).value or cf_sheet.cell(row=row, column=2).value
    value = cf_sheet.cell(row=row, column=48).value  # Col 48 = Q28
    if value:
        print(f"  Row {row}: {label} = €{value:,.2f}")

# 检查Property Disposal at exit (Row 56-59)
print("\n📋 Excel Property Disposal at Q28:")
for row in range(55, 65):
    label = cf_sheet.cell(row=row, column=3).value
    value = cf_sheet.cell(row=row, column=48).value
    if value:
        print(f"  Row {row}: {label} = €{value:,.2f}")

wb.close()
