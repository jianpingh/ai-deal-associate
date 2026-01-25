"""
分析Excel退出价值计算逻辑
"""
import openpyxl
import os

TEMPLATE_PATH = r"d:\work\110agenticAI\code\ai-deal-associate\backend\data\templates\MS Canopy Template -v5.xlsx"

wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)

print("="*70)
print("分析Excel退出价值计算逻辑")
print("="*70)

# 检查Cash Flows表的退出相关单元格
cf_sheet = wb["Cash Flows"]

print("\n📊 Cash Flows Sheet - Exit Values:")
print("-"*50)

# 找到退出季度 (hold_period=5 years = 20 quarters)
exit_col = 'AH'  # Q20 (Column T=Q0, U=Q1, ... AH=Q20)
# T=20, U=21, V=22... 所以 Q20 = T+20 = 列 AH (实际计算)
# T是第20列，Q20是T+20=40列，即AN

# 让我读取实际的列头
print("\n检查季度列...")
for col_idx in range(20, 65):  # T=20 onwards
    cell = cf_sheet.cell(row=78, column=col_idx)  # Date row
    if cell.value:
        print(f"Col {col_idx}: {cell.value}")
        if col_idx == 60:  # Stop early
            break

print("\n📋 关键行检查 (Row 23 - GRI, Row 61 - Purchase Price, Row 102 - Levered CF):")
for row_num in [23, 61, 102]:
    print(f"\nRow {row_num}:")
    for col_idx in range(20, 45):  # T to AS
        cell = cf_sheet.cell(row=row_num, column=col_idx)
        if cell.value is not None:
            print(f"  Col {col_idx}: {cell.value}")

# 检查Exit Value行
print("\n📊 Exit Value相关行:")
exit_rows = [57, 58, 59, 60, 61, 62]  # Property Disposal相关
for row in exit_rows:
    label = cf_sheet.cell(row=row, column=1).value or cf_sheet.cell(row=row, column=2).value
    print(f"\nRow {row}: {label}")
    for col_idx in range(20, 45):
        cell = cf_sheet.cell(row=row, column=col_idx)
        if cell.value is not None:
            print(f"  Col {col_idx}: {cell.value}")

# 检查Money Page的退出计算
mp = wb["Money Page"]
print("\n📊 Money Page - Exit/Returns:")
print("-"*50)

# E47-E55区域
for row in range(47, 60):
    label = mp.cell(row=row, column=3).value or mp.cell(row=row, column=4).value  # C or D
    value = mp.cell(row=row, column=5).value  # E
    if label or value:
        print(f"Row {row}: {label} = {value}")

wb.close()
