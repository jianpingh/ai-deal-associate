import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

print("Starting template creation...")

def create_underwriting_template():
    # Ensure directory exists
    base_dir = os.path.dirname(os.path.abspath(__file__))
    target_dir = os.path.join(base_dir, "data", "templates")
    os.makedirs(target_dir, exist_ok=True)
    
    file_path = os.path.join(target_dir, "financial_model_template.xlsx")
    
    wb = Workbook()
    
    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    currency_format = '#,##0'
    percent_format = '0.00%'
    
    # ==========================================
    # 1. Inputs Sheet
    # ==========================================
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    
    # Headers
    headers = ["Parameter", "Value", "Unit", "Named Range"]
    for col, header in enumerate(headers, 1):
        cell = ws_inputs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Input Data (Default Values)
    input_data = [
        ("Market Rent", 85, "EUR/sqm/yr", "Market_Rent"),
        ("Leasable Area", 10000, "sqm", "Area"),
        ("Rent Growth", 0.03, "%", "Rent_Growth"),
        ("OpEx Ratio", 0.10, "%", "OpEx_Ratio"),
        ("Entry Yield", 0.045, "%", "Entry_Yield"),
        ("Exit Yield", 0.0475, "%", "Exit_Yield"),
        ("LTV", 0.60, "%", "LTV"),
        ("Interest Rate", 0.04, "%", "Interest_Rate"),
        ("Capex", 0, "EUR", "Capex"),
    ]
    
    # Write data and create named ranges
    for row_idx, (param, val, unit, name) in enumerate(input_data, 2):
        ws_inputs.cell(row=row_idx, column=1, value=param).border = thin_border
        
        val_cell = ws_inputs.cell(row=row_idx, column=2, value=val)
        val_cell.border = thin_border
        # Apply formatting based on unit
        if "%" in unit:
            val_cell.number_format = percent_format
        else:
            val_cell.number_format = currency_format
            
        ws_inputs.cell(row=row_idx, column=3, value=unit).border = thin_border
        ws_inputs.cell(row=row_idx, column=4, value=name).border = thin_border
        
        # Create Named Range (Scope: Workbook)
        # Use DefinedName directly to avoid deprecation warnings and double-sheet-name bugs
        dne = DefinedName(name, attr_text=f"'{ws_inputs.title}'!$B${row_idx}")
        wb.defined_names[name] = dne

    # Column widths
    ws_inputs.column_dimensions['A'].width = 20
    ws_inputs.column_dimensions['B'].width = 15
    ws_inputs.column_dimensions['C'].width = 15
    ws_inputs.column_dimensions['D'].width = 20

    # ==========================================
    # 2. Summary Calculations (Hidden helper calculations)
    # ==========================================
    # We'll put some derived values here to make formulas cleaner
    ws_inputs.cell(row=12, column=1, value="Calculated Values").font = Font(bold=True)
    
    ws_inputs.cell(row=13, column=1, value="Initial Gross Income")
    ws_inputs.cell(row=13, column=2, value="=Market_Rent * Area")
    wb.defined_names["Initial_Gross_Income"] = DefinedName("Initial_Gross_Income", attr_text=f"'{ws_inputs.title}'!$B$13")
    
    ws_inputs.cell(row=14, column=1, value="Purchase Price")
    ws_inputs.cell(row=14, column=2, value="=Initial_Gross_Income / Entry_Yield")
    wb.defined_names["Purchase_Price"] = DefinedName("Purchase_Price", attr_text=f"'{ws_inputs.title}'!$B$14")
    
    ws_inputs.cell(row=15, column=1, value="Loan Amount")
    ws_inputs.cell(row=15, column=2, value="=Purchase_Price * LTV")
    wb.defined_names["Loan_Amount"] = DefinedName("Loan_Amount", attr_text=f"'{ws_inputs.title}'!$B$15")
    
    ws_inputs.cell(row=16, column=1, value="Equity Invested")
    ws_inputs.cell(row=16, column=2, value="=Purchase_Price - Loan_Amount + Capex")
    wb.defined_names["Equity_Invested"] = DefinedName("Equity_Invested", attr_text=f"'{ws_inputs.title}'!$B$16")

    # ==========================================
    # 3. Cash Flow Sheet
    # ==========================================
    ws_cf = wb.create_sheet("Cash Flow")
    
    # Headers
    cf_headers = ["Item"] + [f"Year {i}" for i in range(1, 11)]
    for col, header in enumerate(cf_headers, 1):
        cell = ws_cf.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Rows
    # 1. Potential Gross Income
    ws_cf.cell(row=2, column=1, value="Potential Gross Income")
    for year in range(1, 11):
        # Formula: Initial_Gross_Income * (1 + Rent_Growth)^(Year-1)
        col_letter = get_column_letter(year + 1)
        formula = f"=Initial_Gross_Income * (1 + Rent_Growth)^{year-1}"
        cell = ws_cf.cell(row=2, column=year+1, value=formula)
        cell.number_format = currency_format

    # 2. OpEx
    ws_cf.cell(row=3, column=1, value="Operating Expenses")
    for year in range(1, 11):
        col_letter = get_column_letter(year + 1)
        # Formula: -1 * PGI * OpEx_Ratio
        formula = f"=-{col_letter}2 * OpEx_Ratio"
        cell = ws_cf.cell(row=3, column=year+1, value=formula)
        cell.number_format = currency_format

    # 3. NOI
    ws_cf.cell(row=4, column=1, value="Net Operating Income (NOI)").font = Font(bold=True)
    for year in range(1, 11):
        col_letter = get_column_letter(year + 1)
        formula = f"=SUM({col_letter}2:{col_letter}3)"
        cell = ws_cf.cell(row=4, column=year+1, value=formula)
        cell.number_format = currency_format
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    # 4. Debt Service
    ws_cf.cell(row=5, column=1, value="Debt Service")
    for year in range(1, 11):
        # Interest Only for simplicity: Loan_Amount * Interest_Rate
        formula = f"=-Loan_Amount * Interest_Rate"
        cell = ws_cf.cell(row=5, column=year+1, value=formula)
        cell.number_format = currency_format

    # 5. Net Cash Flow
    ws_cf.cell(row=6, column=1, value="Net Cash Flow").font = Font(bold=True)
    for year in range(1, 11):
        col_letter = get_column_letter(year + 1)
        formula = f"={col_letter}4 + {col_letter}5" # NOI + Debt Service (which is negative)
        cell = ws_cf.cell(row=6, column=year+1, value=formula)
        cell.number_format = currency_format
        cell.font = Font(bold=True)

    # 6. Exit Valuation (Year 10)
    ws_cf.cell(row=8, column=1, value="Exit Valuation (Year 10)").font = Font(bold=True)
    
    ws_cf.cell(row=9, column=1, value="Exit NOI (Forward)")
    # Assume Year 11 NOI = Year 10 NOI * (1+Growth)
    ws_cf.cell(row=9, column=2, value="=L4 * (1+Rent_Growth)") # L4 is Year 10 NOI
    ws_cf.cell(row=9, column=2).number_format = currency_format
    
    ws_cf.cell(row=10, column=1, value="Exit Value")
    ws_cf.cell(row=10, column=2, value="=B9 / Exit_Yield")
    ws_cf.cell(row=10, column=2).number_format = currency_format
    
    ws_cf.cell(row=11, column=1, value="Loan Repayment")
    ws_cf.cell(row=11, column=2, value="=-Loan_Amount")
    ws_cf.cell(row=11, column=2).number_format = currency_format
    
    ws_cf.cell(row=12, column=1, value="Net Sale Proceeds")
    ws_cf.cell(row=12, column=2, value="=SUM(B10:B11)")
    ws_cf.cell(row=12, column=2).number_format = currency_format
    ws_cf.cell(row=12, column=2).font = Font(bold=True)

    # 7. Levered Cash Flow Stream
    ws_cf.cell(row=14, column=1, value="Levered Cash Flow Stream").font = Font(bold=True)
    
    # Year 0 (Equity)
    ws_cf.cell(row=14, column=2, value="Year 0")
    ws_cf.cell(row=15, column=2, value="=-Equity_Invested")
    ws_cf.cell(row=15, column=2).number_format = currency_format
    
    # Years 1-10
    for year in range(1, 11):
        col_letter = get_column_letter(year + 2) # Start at C
        ws_cf.cell(row=14, column=year+2, value=f"Year {year}")
        
        # Cash Flow from Ops
        cf_ref = f"{get_column_letter(year+1)}6" # Row 6 is Net Cash Flow
        
        if year == 10:
            # Add Sale Proceeds
            formula = f"={cf_ref} + B12"
        else:
            formula = f"={cf_ref}"
            
        cell = ws_cf.cell(row=15, column=year+2, value=formula)
        cell.number_format = currency_format

    # 8. Returns Metrics
    ws_cf.cell(row=17, column=1, value="Returns Metrics").font = Font(bold=True)
    
    ws_cf.cell(row=18, column=1, value="Levered IRR")
    ws_cf.cell(row=18, column=2, value="=IRR(B15:L15)")
    ws_cf.cell(row=18, column=2).number_format = percent_format
    
    ws_cf.cell(row=19, column=1, value="Equity Multiple")
    ws_cf.cell(row=19, column=2, value="=(SUM(C15:L15) + ABS(B15)) / ABS(B15)")
    ws_cf.cell(row=19, column=2).number_format = "0.00x"

    ws_cf.cell(row=20, column=1, value="Yield on Cost")
    ws_cf.cell(row=20, column=2, value="=C4 / (Purchase_Price + Capex)")
    ws_cf.cell(row=20, column=2).number_format = percent_format

    ws_cf.column_dimensions['A'].width = 25

    # ==========================================
    # 4. Rent Roll Sheet
    # ==========================================
    ws_rr = wb.create_sheet("Rent Roll")
    
    rr_headers = ["Tenant Name", "Unit", "Area (sqm)", "Lease Start", "Lease End", "Annual Rent (EUR)", "Rent/sqm/yr"]
    for col, header in enumerate(rr_headers, 1):
        cell = ws_rr.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws_rr.column_dimensions[get_column_letter(col)].width = 15
        
    ws_rr.column_dimensions['A'].width = 30 # Tenant Name
    ws_rr.column_dimensions['F'].width = 20 # Annual Rent

    print(f"Creating rich template at: {file_path}")
    wb.save(file_path)
    print("Template created successfully.")

if __name__ == "__main__":
    try:
        create_underwriting_template()
    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
