import openpyxl
import os

def fix_excel_template():
    # Path to the template
    base_dir = os.getcwd()
    template_path = os.path.join(base_dir, "backend", "data", "templates", "financial_model_template.xlsx")
    
    print(f"Loading template from: {template_path}")
    
    if not os.path.exists(template_path):
        print("Error: Template file not found!")
        return

    try:
        wb = openpyxl.load_workbook(template_path)
        
        if "Cash Flow" in wb.sheetnames:
            ws = wb["Cash Flow"]
            print("Switched to 'Cash Flow' sheet.")
        else:
            ws = wb.active
            print(f"Using active sheet: {ws.title}")
        
        # Fix Yield on Cost Formula
        # Target: Row 20, Column 2 (B20)
        # Formula: =B4 / (Purchase_Price + Capex)
        
        # Verify if Row 20 is indeed Yield on Cost
        label_cell = ws.cell(row=20, column=1).value
        print(f"Row 20 Label: {label_cell}")
        
        # Update the formula
        target_cell = ws.cell(row=20, column=2)
        new_formula = "=B4/(Purchase_Price+Capex)"
        target_cell.value = new_formula
        target_cell.number_format = '0.00%'
        
        print(f"Updated Yield on Cost formula in {target_cell.coordinate} to: {new_formula}")
        
        wb.save(template_path)
        print("Template successfully updated and saved.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    fix_excel_template()