from langchain_core.tools import tool
import openpyxl
import pandas as pd
from typing import Dict, Any, List

@tool
def read_excel_sheet(file_path: str, sheet_name: str = None) -> str:
    """
    Read data from an Excel file.
    Useful for extracting financial data, comps, or other structured information.
    
    Args:
        file_path: The absolute path to the Excel file.
        sheet_name: The name of the sheet to read. If None, reads the first sheet.
    """
    try:
        # Using pandas for easier data extraction and formatting
        df = pd.read_excel(file_path, sheet_name=sheet_name if sheet_name else 0)
        return df.to_markdown(index=False)
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

@tool
def update_financial_model(file_path: str, updates: Dict[str, Any]) -> str:
    """
    Update specific cells in a financial model Excel file.
    Useful for applying new assumptions to a DCF or LBO model.
    
    Args:
        file_path: The absolute path to the Excel file.
        updates: A dictionary where keys are cell references (e.g., 'B4', 'Sheet1!C10') 
                 and values are the new values to write.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        updated_cells = []
        for cell_ref, value in updates.items():
            # Handle sheet specification (e.g., "Assumptions!B5")
            if "!" in cell_ref:
                sheet_name, cell_addr = cell_ref.split("!")
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws[cell_addr] = value
                    updated_cells.append(f"{sheet_name}!{cell_addr}={value}")
            else:
                # Default to active sheet if no sheet specified
                ws = wb.active
                ws[cell_ref] = value
                updated_cells.append(f"{cell_ref}={value}")
        
        wb.save(file_path)
        return f"Successfully updated cells: {', '.join(updated_cells)}"
    except Exception as e:
        return f"Error updating financial model: {str(e)}"
