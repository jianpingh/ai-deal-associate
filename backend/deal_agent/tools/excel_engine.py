from langchain_core.tools import tool
import openpyxl
import pandas as pd
from typing import Dict, Any, List

@tool
def read_excel_sheet(file_path: str, sheet_name: str = None) -> str:
    """
    Read data from an Excel file.
    Useful for extracting financial data, comps, or other structured information.
    
    Args:
        file_path: The absolute path to the Excel file.
        sheet_name: The name of the sheet to read. If None, reads the first sheet.
    """
    try:
        # Using pandas for easier data extraction and formatting
        df = pd.read_excel(file_path, sheet_name=sheet_name if sheet_name else 0)
        return df.to_markdown(index=False)
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

@tool
def update_financial_model(file_path: str, updates: Dict[str, Any]) -> str:
    """
    Update specific cells in a financial model Excel file.
    Useful for applying new assumptions to a DCF or LBO model.
    
    Args:
        file_path: The absolute path to the Excel file.
        updates: A dictionary where keys are cell references (e.g., 'B4', 'Sheet1!C10') 
                 and values are the new values to write.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        updated_cells = []
        for cell_ref, value in updates.items():
            # Handle sheet specification (e.g., "Assumptions!B5")
            if "!" in cell_ref:
                sheet_name, cell_addr = cell_ref.split("!")
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws[cell_addr] = value
                    updated_cells.append(f"{sheet_name}!{cell_addr}={value}")
            else:
                # Default to active sheet if no sheet specified
                ws = wb.active
                ws[cell_ref] = value
                updated_cells.append(f"{cell_ref}={value}")
        
        wb.save(file_path)
        return f"Successfully updated cells: {', '.join(updated_cells)}"
    except Exception as e:
        return f"Error updating financial model: {str(e)}"

@tool
def fill_excel_named_ranges(file_path: str, data: Dict[str, Any]) -> str:
    """
    Fill named ranges in an Excel file with provided values.
    
    Args:
        file_path: The absolute path to the Excel file.
        data: A dictionary where keys are named ranges and values are the values to write.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        updated_ranges = []
        
        for name, value in data.items():
            if name in wb.defined_names:
                dne = wb.defined_names[name]
                
                # destinations yields (sheet_title, coord)
                try:
                    dests = list(dne.destinations)
                    for sheet_title, coord in dests:
                        ws = wb[sheet_title]
                        # Handle range coordinates like $A$1 or $A$1:$B$2
                        # Remove $ signs for cleaner handling if needed, but openpyxl handles them usually
                        if ':' in coord:
                            # If it's a range, set the top-left cell
                            top_left = coord.split(':')[0]
                            ws[top_left] = value
                        else:
                            ws[coord] = value
                    updated_ranges.append(f"{name}")
                except Exception as name_err:
                    return f"Could not resolve named range {name}: {str(name_err)}"
            else:
                return f"Named range '{name}' not found in workbook."
        
        wb.save(file_path)
        return f"Successfully updated named ranges: {', '.join(updated_ranges)}"
    except Exception as e:
        return f"Error updating named ranges: {str(e)}"

@tool
def write_list_to_excel(file_path: str, sheet_name: str, data: List[List[Any]], start_row: int = 2, start_col: int = 1) -> str:
    """
    Write a list of lists (table data) to a specific sheet in an Excel file.
    Useful for writing rent rolls, comps lists, or schedules.
    
    Args:
        file_path: Absolute path to the Excel file.
        sheet_name: Name of the sheet to write to.
        data: List of rows, where each row is a list of values.
        start_row: Row number to start writing (1-based). Default is 2 (assuming header is row 1).
        start_col: Column number to start writing (1-based). Default is 1 (Column A).
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return f"Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        
        for r_idx, row_data in enumerate(data):
            for c_idx, value in enumerate(row_data):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
                
        wb.save(file_path)
        return f"Successfully wrote {len(data)} rows to sheet '{sheet_name}'."
    except Exception as e:
        return f"Error writing list to Excel: {str(e)}"
