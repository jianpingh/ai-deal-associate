"""
分析Excel退出季度现金流
"""
import openpyxl
import os

TEMPLATE_PATH = r"d:\work\110agenticAI\code\ai-deal-associate\backend\data\templates\MS Canopy Template -v5.xlsx"

wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
cf_sheet = wb["Cash Flows"]

print("="*70)
print("分析Excel退出季度现金流")
print("="*70)

# 找到5年退出季度
# Q0 = Col T (col_idx=20), Start date 2025-09-30
# Q20 = Col T+20 = Col AN (col_idx=40)
# 5年后 = 2030-09-30

print("\n📊 Row 102 (Levered CF) - 所有非零值:")
exit_found = False
for col_idx in range(20, 65):
    cell = cf_sheet.cell(row=102, column=col_idx)
    date_cell = cf_sheet.cell(row=78, column=col_idx)
    if cell.value is not None and cell.value != 0:
        print(f"Col {col_idx} ({date_cell.value}): €{cell.value:,.2f}")
        if col_idx == 40:
            print("   ^^^ This is Q20 (5-year exit)")
            exit_found = True

if not exit_found:
    print("\n检查Q20 (Col 40)的详细情况:")
    for row in [23, 56, 57, 58, 59, 61, 97, 99, 100, 101, 102]:
        label_cell = cf_sheet.cell(row=row, column=2) or cf_sheet.cell(row=row, column=3)
        value = cf_sheet.cell(row=40, column=row)
        actual_value = cf_sheet.cell(row=row, column=40).value
        print(f"Row {row}: {cf_sheet.cell(row=row, column=2).value or cf_sheet.cell(row=row, column=3).value} = {actual_value}")

# 检查退出价值计算 (Row 56-59)
print("\n📊 Property Disposal rows at Q20 (Col 40):")
for row in range(55, 65):
    label = cf_sheet.cell(row=row, column=3).value or cf_sheet.cell(row=row, column=2).value
    value = cf_sheet.cell(row=row, column=40).value
    print(f"Row {row}: {label} = {value}")

# 检查Exit Cap和Exit NOI
print("\n📊 检查Exit相关参数:")
# Input Other - Exit Cap Rate
input_other = wb["Input Other"]
print(f"Exit Cap Rate (Input Other B15): {input_other['B15'].value}")

# 检查Cash Flows Row 56 (Property Disposal) 公式
print("\n📊 检查所有Property Disposal相关行:")
for row in range(55, 65):
    for col in [40, 41, 42]:  # Q20, Q21, Q22
        cell = cf_sheet.cell(row=row, column=col)
        label = cf_sheet.cell(row=row, column=3).value
        print(f"Row {row} Col {col} ({label}): {cell.value}")

# 检查EBITDA at Exit
print("\n📊 检查EBITDA (Row 23) at exit quarters:")
for col in range(38, 45):  # Around Q20
    date = cf_sheet.cell(row=78, column=col).value
    gri = cf_sheet.cell(row=23, column=col).value
    print(f"Col {col} ({date}): GRI = {gri}")

# 计算Excel的隐含退出价值
print("\n📊 计算隐含退出价值:")
# Exit NOI (年化) = 4 quarters of GRI at exit
q20_gri = cf_sheet.cell(row=23, column=40).value  # Col 40 is Q20
if q20_gri:
    print(f"Q20 GRI: €{q20_gri:,.2f}")
    annual_exit_noi = q20_gri * 4
    print(f"Annual Exit NOI (Q20 GRI × 4): €{annual_exit_noi:,.2f}")
    exit_cap = 0.065
    gross_exit_value = annual_exit_noi / exit_cap
    print(f"Gross Exit Value (NOI / Exit Cap): €{gross_exit_value:,.2f}")

wb.close()
