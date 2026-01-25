"""
直接读取Excel Q1的税收信息
"""
import openpyxl

TEMPLATE_PATH = r"d:\work\110agenticAI\code\ai-deal-associate\backend\data\templates\MS Canopy Template -v5.xlsx"

try:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    
    # Input Other - Tax Rate
    input_other = wb["Input Other"]
    print("Input Other Tax-related cells:")
    for row in range(15, 25):
        cell_b = input_other.cell(row=row, column=2).value
        cell_a = input_other.cell(row=row, column=1).value
        if cell_a or cell_b:
            print(f"Row {row}: A={cell_a}, B={cell_b}")
    
    # Money Page - Tax related
    mp = wb["Money Page"]
    print("\nMoney Page Tax-related:")
    for row in range(1, 80):
        for col in [3, 4, 5]:
            cell = mp.cell(row=row, column=col).value
            if cell and 'tax' in str(cell).lower():
                print(f"Row {row}: {cell}")
    
    wb.close()
except Exception as e:
    print(f"Error: {e}")
